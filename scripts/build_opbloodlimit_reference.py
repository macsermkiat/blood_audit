"""Build the OPBloodLimit MSBOS reference CSVs for the pre-op blood-reservation arm.

Reads the four specialty sheets of ``OPBloodLimit.xlsx`` (the hospital's Maximum
Surgical Blood Ordering Schedule) and emits two CSVs beside the source:

  * ``OPBloodLimit.csv``          one row per operation; ICD-9 codes joined with "; "
  * ``OPBloodLimit_by_icd9.csv``  exploded to one row per (operation, ICD-9 code),
                                  keyed on the dotless ``icd9_code_nodot`` for
                                  joining against ``ICD9CM.csv`` / the operative
                                  tables (which store codes dotless, e.g. ``0602``).

The raw ``.xlsx`` is treated as an immutable hospital export: source quirks and the
two known data-entry errors are corrected *here*, in code, with an audit trail --
never by mutating the workbook.

Source quirks handled:
  * merged procedure-group cells (forward-fill down col A)
  * the extra Thai header row in the ``Sx`` sheet
  * ICD-9 codes spread across many trailing columns
  * ICD-9 codes stored as floats (OB-Gyn) with spurious trailing zeros
  * two Ortho rows where Excel mis-parsed a typed "1-2" unit range into a date

Operations with no ICD-9 code in the source are KEPT in the exploded file with
blank code columns (they cannot be matched by code -- the arm must match them by
operation description). ``ICD9CM.csv`` coverage is reported for information only
and is never a reason to drop a row: that master may itself be incomplete.

Run:  uv run python scripts/build_opbloodlimit_reference.py
"""

from __future__ import annotations

import argparse
import csv
import datetime
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

SPECIALTY = {
    "Sx": "General Surgery",
    "Ortho": "Orthopedics",
    "ENT": "ENT (Otolaryngology)",
    "OB-Gyn": "Obstetrics & Gynecology",
}
MEANING = {
    "T/S": "Type and Screen",
    "G/M": "Group and Match (crossmatch)",
    "none": "No blood preparation",
}
HEADER_OPS = {"OPERATION", "หัตถการ"}  # English + the extra Sx Thai header row

# Verified data-entry corrections, applied to the dotted ICD-9 code as written in
# the workbook. Each target is a *structurally invalid* code (not merely absent
# from the ICD-9-CM master), cross-checked against ICD9CM.csv:
#   06.40 -> 06.4   No 4-digit 06.40 exists; 06.4 = "Complete thyroidectomy", and
#                   the ENT sheet already writes this same code correctly as 06.4.
#   4.73  -> 04.73  ICD-9 procedure codes always carry two leading digits; the
#                   leading zero was dropped. 04.73 = "Accessory-hypoglossal
#                   anastomosis" -- exactly the ENT facial-reanimation procedure.
ICD_CORRECTIONS = {"06.40": "06.4", "4.73": "04.73"}

GROUPED_COLS = [
    "specialty", "sheet", "procedure_group", "operation",
    "msbos", "msbos_meaning", "recommended_units", "icd9_codes",
]
EXPLODED_COLS = [
    "icd9_code", "icd9_code_nodot", "specialty", "sheet",
    "procedure_group", "operation", "msbos", "msbos_meaning", "recommended_units",
]


def fmt_icd(v: object) -> str:
    """Render a raw ICD-9 cell to a dotted code string.

    Floats (OB-Gyn stores codes numerically) are capped at two decimals -- the
    ICD-9 maximum -- then stripped of spurious trailing zeros so a genuine
    3-digit code like 74.1 does not masquerade as the 4-digit 74.10.
    """
    if isinstance(v, bool):
        return ""
    if isinstance(v, float):
        return ("%.2f" % v).rstrip("0").rstrip(".")
    if isinstance(v, int):
        return str(v)
    return str(v).strip()


def fmt_units(v: object) -> str:
    """Render the (unlabeled) recommended-units cell."""
    if v is None:
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        # Excel mis-parsed a typed "N-M" unit range into a date (day=N, month=M).
        return f"{v.day}-{v.month}"
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else str(v)
    return str(v).strip()


def norm_msbos(v: object) -> str:
    """Normalise the MSBOS token; fold the 'None'/'none' variants to 'none'."""
    if v is None:
        return ""
    s = str(v).strip()
    return "none" if s.lower() == "none" else s


def parse(src: Path, applied: dict[str, int]) -> list[dict[str, object]]:
    """Parse every sheet into per-operation records with de-duped dotted codes.

    ``applied`` is mutated to count how many times each ICD_CORRECTIONS key fired,
    so a correction that no longer matches the source (fixed upstream) is visible.
    """
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    ops: list[dict[str, object]] = []
    for sheet in wb.sheetnames:
        group = ""
        for r in wb[sheet].iter_rows(min_row=1, values_only=True):
            op = r[1]
            if op is None or str(op).strip() == "" or str(op).strip() in HEADER_OPS:
                continue
            if r[0] is not None and str(r[0]).strip() != "":
                group = str(r[0]).strip()
            codes: list[str] = []
            for c in range(4, len(r)):
                if r[c] is not None and str(r[c]).strip() != "":
                    dc = fmt_icd(r[c])
                    if dc in ICD_CORRECTIONS:
                        applied[dc] = applied.get(dc, 0) + 1
                        dc = ICD_CORRECTIONS[dc]
                    if dc and dc not in codes:
                        codes.append(dc)
            msbos = norm_msbos(r[2])
            ops.append({
                "specialty": SPECIALTY.get(sheet, sheet),
                "sheet": sheet,
                "procedure_group": group,
                "operation": str(op).strip(),
                "msbos": msbos,
                "msbos_meaning": MEANING.get(msbos, ""),
                "recommended_units": fmt_units(r[3]),
                "codes": codes,
            })
    return ops


def write_grouped(ops: list[dict[str, object]], out: Path) -> None:
    with out.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=GROUPED_COLS)
        w.writeheader()
        for o in ops:
            row = {k: o[k] for k in GROUPED_COLS if k != "icd9_codes"}
            row["icd9_codes"] = "; ".join(o["codes"])  # type: ignore[arg-type]
            w.writerow(row)


def build_exploded(ops: list[dict[str, object]]) -> list[dict[str, str]]:
    """Explode to one row per (operation, code); no-ICD operations kept with blanks."""
    rows: list[dict[str, str]] = []
    for o in ops:
        base: dict[str, str] = {
            "specialty": str(o["specialty"]), "sheet": str(o["sheet"]),
            "procedure_group": str(o["procedure_group"]), "operation": str(o["operation"]),
            "msbos": str(o["msbos"]), "msbos_meaning": str(o["msbos_meaning"]),
            "recommended_units": str(o["recommended_units"]),
        }
        codes: list[str] = o["codes"]  # type: ignore[assignment]
        if codes:
            for dc in codes:
                rows.append({"icd9_code": dc, "icd9_code_nodot": dc.replace(".", ""), **base})
        else:
            rows.append({"icd9_code": "", "icd9_code_nodot": "", **base})
    # Coded rows first (sorted by dotless key), blank-code rows last.
    rows.sort(key=lambda x: (x["icd9_code_nodot"] == "", x["icd9_code_nodot"], x["operation"]))
    return rows


def write_exploded(rows: list[dict[str, str]], out: Path) -> None:
    with out.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=EXPLODED_COLS)
        w.writeheader()
        w.writerows(rows)


def load_ref_codes(ref: Path) -> set[str]:
    codes: set[str] = set()
    with ref.open(encoding="utf-8") as f:
        for row in csv.DictReader(f):
            c = (row.get("Icd9cm") or "").strip()
            if c:
                codes.add(c)
    return codes


def report(
    ops: list[dict[str, object]],
    exploded: list[dict[str, str]],
    ref_codes: set[str],
    applied: dict[str, int],
) -> None:
    coded = [x for x in exploded if x["icd9_code_nodot"]]
    blank = [x for x in exploded if not x["icd9_code_nodot"]]
    distinct = {x["icd9_code_nodot"] for x in coded}

    # correction audit (fail loud if a mapping went stale)
    for key in ICD_CORRECTIONS:
        n = applied.get(key, 0)
        tag = "applied" if n else "STALE (no longer in source -- consider removing)"
        print(f"[correction] {key} -> {ICD_CORRECTIONS[key]}: {n}x {tag}", file=sys.stderr)

    # coverage: informational only -- never drops a row
    unmatched = sorted({(x["icd9_code"], x["icd9_code_nodot"]) for x in coded
                        if x["icd9_code_nodot"] not in ref_codes})
    print(f"[coverage] codes absent from ICD9CM.csv (kept anyway): {len(unmatched)}",
          file=sys.stderr)
    for dotted, nodot in unmatched:
        print(f"    KEEP {nodot} ({dotted})", file=sys.stderr)

    # same code -> conflicting recommendation (arm must disambiguate by operation)
    bycode: dict[str, set[tuple[str, str]]] = defaultdict(set)
    for x in coded:
        bycode[x["icd9_code_nodot"]].add((x["msbos"], x["recommended_units"]))
    conflicts = {k: v for k, v in bycode.items() if len(v) > 1}

    print(f"[blank]    operations with no ICD-9 code (kept, match by name): {len(blank)}",
          file=sys.stderr)
    for x in blank:
        print(f"    {x['sheet']}: {x['operation']}", file=sys.stderr)
    print(f"[conflicts] codes with >1 (msbos,units): {len(conflicts)} of {len(distinct)} distinct",
          file=sys.stderr)
    print(f"grouped rows : {len(ops)}", file=sys.stderr)
    print(f"exploded rows: {len(exploded)} ({len(coded)} coded + {len(blank)} blank)",
          file=sys.stderr)


def main() -> None:
    default_raw = Path(__file__).resolve().parents[1].parent / "Bloodbank" / "data" / "raw"
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    ap.add_argument(
        "--raw-dir", type=Path, default=default_raw,
        help="Folder holding OPBloodLimit.xlsx + ICD9CM.csv, and where the CSVs "
             "are written (default: Bloodbank/data/raw).",
    )
    ap.add_argument(
        "--out-dir", type=Path, default=None,
        help="Where to write the CSVs (default: same as --raw-dir).",
    )
    args = ap.parse_args()

    raw_dir: Path = args.raw_dir
    out_dir: Path = args.out_dir or raw_dir
    src = raw_dir / "OPBloodLimit.xlsx"
    ref = raw_dir / "ICD9CM.csv"
    if not src.is_file():
        sys.stderr.write(f"ERROR: source workbook not found: {src}\n")
        sys.exit(1)

    applied: dict[str, int] = {}
    ops = parse(src, applied)
    write_grouped(ops, out_dir / "OPBloodLimit.csv")
    exploded = build_exploded(ops)
    write_exploded(exploded, out_dir / "OPBloodLimit_by_icd9.csv")

    ref_codes = load_ref_codes(ref) if ref.is_file() else set()
    if not ref_codes:
        sys.stderr.write(f"WARN: ICD9CM.csv not found at {ref}; coverage check skipped\n")
    report(ops, exploded, ref_codes, applied)
    print(f"wrote {out_dir / 'OPBloodLimit.csv'} and {out_dir / 'OPBloodLimit_by_icd9.csv'}")


if __name__ == "__main__":
    main()
